#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/5/26 10:18
# @Author  : 刘宇
# @File    : ExcelUtil.py
import os

from openpyxl import load_workbook


def excel2dict(filePath, sheetName='Sheet1', skipRow=0):
    """
    该函数用于读取excel指定的sheetName中的表格内容
    :param filePath: 提供需要读取的excel表格路径
    :param sheetName: 提供需要读取的SheetName名称
    :return: 数据字典
    """
    if not os.path.isfile(filePath):
        print(f"该{filePath}不存在。")
        return {}
    wb = load_workbook(filePath)
    try:
        ws = wb[sheetName]
    except Exception as e:
        print(f"该{sheetName}不存在。{e}")
        return {}
    finally:
        pass
    data = []
    for row in ws.iter_rows(min_row=1 + skipRow):
        data.append([cell.value for cell in row])
    data = list(map(list, zip(*data)))
    dict = {}
    for row in data:
        dict[row[0]] = row[1:]
    return dict


if __name__ == "__main__":
    print(excel2dict(filePath='../data/python.xlsx', sheetName='Sheet1', skipRow=1))
    run_code = 0
